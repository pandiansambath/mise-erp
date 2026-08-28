"""Downloading a supplier's price list.

"for each vendor I need one download feature — I can download vendor items or
items with price details."

He asked the assistant for this first and was told it could not be seen. That
was TRUE, and not the model's fault: there was no way to get it out of the app
at all. The tenant test below is the one that matters — the vendor id is in the
URL, so a supplier id from another restaurant must not return an item list.
"""

from decimal import Decimal
from io import BytesIO

import pytest

from app.hotels.models import Hotel
from app.inventory import service as inv
from app.vendors import service as ven


async def _supplier_with_a_price(db, hotel_id, *, vendor="RUDRA EXIM", item="Basmati Rice"):
    it = await inv.create_item(db, hotel_id, name=item, unit="kg")
    v = await ven.create_vendor(db, hotel_id, name=vendor)
    await ven.upsert_vendor_item(db, v.id, it.id, Decimal("5.40"))
    return v, it


@pytest.mark.asyncio
async def test_the_rows_carry_the_item_and_the_price(db, hotel):
    v, _ = await _supplier_with_a_price(db, hotel.id)

    rows = await ven.vendor_price_rows(db, v.id, hotel.id)

    assert len(rows) == 1
    assert rows[0]["item"] == "Basmati Rice"
    assert rows[0]["unit"] == "kg"
    assert Decimal(str(rows[0]["price"])) == Decimal("5.40")


@pytest.mark.asyncio
async def test_another_restaurants_items_are_not_returned(db, hotel):
    """The vendor id comes from the URL. Scoping only on it would leak."""
    other = Hotel(name="Someone Else", country="GB", base_currency="GBP", city="Leeds")
    db.add(other)
    await db.commit()
    await db.refresh(other)

    theirs, _ = await _supplier_with_a_price(db, other.id, vendor="Their Supplier")

    # Ask for THEIR vendor while claiming to be OUR hotel.
    rows = await ven.vendor_price_rows(db, theirs.id, hotel.id)

    assert rows == [], "a vendor id from another tenant returned their item list"


@pytest.mark.asyncio
async def test_a_supplier_with_nothing_listed_downloads_an_empty_sheet(db, hotel):
    v = await ven.create_vendor(db, hotel.id, name="New Supplier")
    assert await ven.vendor_price_rows(db, v.id, hotel.id) == []


@pytest.mark.asyncio
async def test_the_download_is_a_real_spreadsheet(client, db, hotel, make_user, auth_header):
    v, _ = await _supplier_with_a_price(db, hotel.id)
    owner = await make_user("buyer@x.com", "SUPER_ADMIN")

    r = await client.get(f"/api/vendors/{v.id}/price-list.xlsx", headers=auth_header(owner))

    assert r.status_code == 200, r.text
    assert "attachment" in r.headers.get("content-disposition", "")

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.content))
    text = " ".join(
        str(c.value) for row in wb.active.iter_rows() for c in row if c.value is not None
    )
    assert "Basmati Rice" in text
    assert "5.4" in text or "5.40" in text


@pytest.mark.asyncio
async def test_general_staff_cannot_download_a_price_list(client, db, hotel, make_user, auth_header):
    v, _ = await _supplier_with_a_price(db, hotel.id)
    floor = await make_user("floor2@x.com", "STAFF")

    r = await client.get(f"/api/vendors/{v.id}/price-list.xlsx", headers=auth_header(floor))

    assert r.status_code == 403
