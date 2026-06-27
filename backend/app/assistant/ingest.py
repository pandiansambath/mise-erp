"""Document onboarding — upload a PDF/image/CSV of your existing items or
suppliers and the Copilot reads it (Gemini multimodal), extracts structured
rows, and (after you confirm) bulk-creates them via the normal services.

Two steps so a human always confirms before anything is written:
  • extract()  — read the file, return proposed rows. Writes NOTHING.
  • commit()   — create the confirmed rows, scoped to the user's hotel + RBAC,
                 audit-logged so it's traceable/undoable.
"""
from __future__ import annotations

import base64
import json
from datetime import date as date_type
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.audit import service as audit
from app.auth.models import User
from app.core.config import settings
from app.core.rbac import has_permission
from app.employees import service as employee_service
from app.inventory import service as inventory_service
from app.recipes import service as recipe_service
from app.sales import service as sales_service
from app.vendors import service as vendor_service

from .provider import ProviderError, is_configured, post_gemini

_ENDPOINT = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"

_ITEM_SCHEMA = {
    "type": "array",
    "items": {
        "type": "object",
        "properties": {
            "name": {"type": "string"},
            "unit": {"type": "string"},
            "category": {"type": "string"},
            "current_stock": {"type": "number"},
            "cost_price": {"type": "number"},
        },
        "required": ["name", "unit"],
    },
}
_VENDOR_SCHEMA = {
    "type": "array",
    "items": {
        "type": "object",
        "properties": {
            "name": {"type": "string"},
            "category": {"type": "string"},
            "contact_person": {"type": "string"},
            "mobile": {"type": "string"},
            "email": {"type": "string"},
        },
        "required": ["name"],
    },
}

_RECIPE_SCHEMA = {
    "type": "array",
    "items": {
        "type": "object",
        "properties": {
            "name": {"type": "string"},
            "category": {"type": "string"},
            "selling_price": {"type": "number"},
        },
        "required": ["name"],
    },
}
_EMPLOYEE_SCHEMA = {
    "type": "array",
    "items": {
        "type": "object",
        "properties": {
            "name": {"type": "string"},
            "job_title": {"type": "string"},
            "monthly_salary": {"type": "number"},
            "hourly_rate": {"type": "number"},
            "mobile": {"type": "string"},
        },
        "required": ["name"],
    },
}
_SALES_SCHEMA = {
    "type": "array",
    "items": {
        "type": "object",
        "properties": {
            "date": {"type": "string"},
            "channel": {"type": "string"},
            "amount": {"type": "number"},
        },
        "required": ["amount"],
    },
}

# kind -> how to extract + how to write it
KINDS: dict[str, dict] = {
    "items": {
        "perm": "inventory:write",
        "label": "stock items",
        "schema": _ITEM_SCHEMA,
        "str_fields": ["name", "unit", "category"],
        "num_fields": ["current_stock", "cost_price"],
        "prompt": (
            "You are reading a restaurant's stock / inventory document. Extract EVERY "
            "distinct stock item. For each: name (the ingredient or product), unit "
            "(kg, g, l, ml, each, pack, bottle…), category if shown (e.g. Vegetables, "
            "Dairy, Meat, Dry Goods, Packaging), current_stock as a number if a quantity "
            "is shown, and cost_price per unit as a number if a price is shown. Skip "
            "headers, totals and section titles. Omit any field that isn't present."
        ),
    },
    "vendors": {
        "perm": "vendors:write",
        "label": "suppliers",
        "schema": _VENDOR_SCHEMA,
        "str_fields": ["name", "category", "contact_person", "mobile", "email"],
        "num_fields": [],
        "prompt": (
            "You are reading a restaurant's supplier / vendor list. Extract EVERY "
            "supplier. For each: name (company or person), category (what they supply, "
            "e.g. Vegetables, Meat, Dairy, Packaging), contact_person, mobile (phone "
            "number), email. Skip anything that isn't a supplier. Omit absent fields."
        ),
    },
    "recipes": {
        "perm": "recipes:write",
        "label": "dishes",
        "schema": _RECIPE_SCHEMA,
        "str_fields": ["name", "category"],
        "num_fields": ["selling_price"],
        "prompt": (
            "You are reading a restaurant's MENU or recipe list. Extract EVERY dish. "
            "For each: name (the dish), category if shown (e.g. Starters, Mains, Breads, "
            "Rice, Desserts, Drinks), and selling_price as a number if a menu price is "
            "shown. Skip section headers, prices-only lines and notes. Omit absent fields."
        ),
    },
    "employees": {
        "perm": "employees:write",
        "label": "staff",
        "schema": _EMPLOYEE_SCHEMA,
        "str_fields": ["name", "job_title", "mobile"],
        "num_fields": ["monthly_salary", "hourly_rate"],
        "prompt": (
            "You are reading a restaurant's STAFF / employee list. Extract EVERY person. "
            "For each: name (full name), job_title (e.g. Chef, Waiter, Cashier, Manager, "
            "Kitchen Porter), monthly_salary as a number if shown, hourly_rate as a number "
            "if shown, mobile (phone number). Skip headers and totals. Omit absent fields."
        ),
    },
    "sales": {
        "perm": "sales:write",
        "label": "sales entries",
        "schema": _SALES_SCHEMA,
        "str_fields": ["date", "channel"],
        "num_fields": ["amount"],
        "prompt": (
            "You are reading a restaurant's past SALES / takings / revenue report. Extract "
            "the takings as rows: date (the day, any format shown), channel (e.g. Dine-in, "
            "Takeaway, Uber Eats, Deliveroo, Just Eat — use 'Dine-in' if not specified), and "
            "amount (the takings for that day/channel, as a number). One row per day per "
            "channel. Skip totals, subtotals, headers and any non-sales lines."
        ),
    },
}

MAX_BYTES = 15 * 1024 * 1024  # Gemini inline-data ceiling is ~20MB; stay under it


def kind_perm(kind: str) -> str | None:
    cfg = KINDS.get(kind)
    return cfg["perm"] if cfg else None


_EXCEL_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls (older; openpyxl reads xlsx, best-effort here)
}


def _is_excel(mime: str) -> bool:
    return (mime or "").split(";")[0].strip() in _EXCEL_MIMES


def _xlsx_to_csv(file_bytes: bytes, max_rows: int = 500) -> str:
    """Flatten the first worksheet to CSV text so the model can read it. Best-effort:
    if openpyxl/the file fails, returns '' (the model then sees an empty sheet)."""
    import csv
    import io

    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        out = io.StringIO()
        w = csv.writer(out)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            w.writerow(["" if c is None else c for c in row])
        return out.getvalue()
    except Exception:  # noqa: BLE001 — bad/unsupported file → let the model see nothing
        return ""


async def extract(file_bytes: bytes, mime: str, kind: str) -> list[dict]:
    """Read the uploaded document and return proposed rows (writes nothing)."""
    if kind not in KINDS:
        raise ValueError(f"Unknown document kind '{kind}'")
    if not is_configured():
        raise ProviderError("no api key")
    cfg = KINDS[kind]

    try:
        import httpx
    except ImportError as exc:
        raise ProviderError("httpx not installed") from exc

    url = _ENDPOINT.format(model=settings.assistant_model)
    # Gemini can't read an .xlsx binary, so convert spreadsheets to CSV text first.
    if _is_excel(mime):
        sheet = _xlsx_to_csv(file_bytes)
        parts = [{"text": cfg["prompt"] + "\n\nSPREADSHEET CONTENTS (CSV):\n" + sheet}]
    else:
        parts = [
            {"text": cfg["prompt"]},
            {"inline_data": {"mime_type": mime, "data": base64.b64encode(file_bytes).decode()}},
        ]
    body = {
        "contents": [{"role": "user", "parts": parts}],
        "generationConfig": {
            "responseMimeType": "application/json",
            "responseSchema": cfg["schema"],
            "temperature": 0.1,
            "maxOutputTokens": 8192,
        },
    }
    try:
        async with httpx.AsyncClient(timeout=120) as client:
            resp = await post_gemini(client, url, body)  # rotates keys on 429
            if resp.status_code >= 300:
                raise ProviderError(f"gemini {resp.status_code}: {resp.text[:300]}")
            data = resp.json()
        text = "".join(
            p.get("text", "") for p in data["candidates"][0]["content"]["parts"]
        )
        rows = json.loads(text)
        if not isinstance(rows, list):
            return []
        key = "amount" if kind == "sales" else "name"
        return [r for r in rows if isinstance(r, dict) and r.get(key) is not None]
    except ProviderError:
        raise
    except Exception as exc:  # noqa: BLE001 — network/JSON → caller decides
        raise ProviderError(str(exc)) from exc


def _dec(v: Any) -> str | None:
    try:
        return str(Decimal(str(v)))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _resolve_date(v: Any) -> date_type:
    """Parse a date cell from a report. UK-first (DD/MM). Falls back to today."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date_type):
        return v
    s = str(v or "").strip()
    if not s or s.lower() in ("today", "now"):
        return date_type.today()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d %b %Y", "%d %B %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:11].strip(), fmt).date()
        except ValueError:
            continue
    return date_type.today()


async def _commit_sales(db: AsyncSession, user: User, rows: list[dict]) -> dict:
    """Bulk-import past takings: each row {date, channel, amount} → resolve/create the
    channel, upsert that day, add a line. Date-keyed (no 'name'), so it has its own path."""
    hotel = user.hotel_id
    created: list[str] = []
    skipped: list[str] = []
    for row in rows:
        amt = _dec(row.get("amount"))
        if amt is None or Decimal(amt) <= 0:
            continue
        d = _resolve_date(row.get("date"))
        ch_name = (row.get("channel") or "").strip() or "Dine-in"
        try:
            ch = await sales_service.get_channel_by_name(db, hotel, ch_name)
            if ch is None:
                ch = await sales_service.create_channel(db, hotel, ch_name, Decimal("0"))
            day = await sales_service.upsert_day(db, hotel, d, entered_by=user.id)
            await sales_service.add_line(db, day, ch.id, Decimal(amt), "CARD")
            created.append(f"{d} {ch_name}: {amt}")
        except Exception:  # noqa: BLE001 — bad row → skip, keep going
            skipped.append(str(row.get("date")))
    if created:
        await audit.record(
            db, hotel_id=hotel, user=user, action="assistant.onboard.sales",
            summary=f"Copilot onboarding imported {len(created)} sales entries",
            entity_type="sales",
        )
    return {"kind": "sales", "created": created, "skipped": skipped}


async def commit(db: AsyncSession, user: User, kind: str, rows: list[dict]) -> dict:
    """Create the confirmed rows. Skips duplicates/invalid; audit-logged."""
    cfg = KINDS.get(kind)
    if not cfg:
        return {"error": f"Unknown document kind '{kind}'"}
    if not has_permission(user.role, cfg["perm"]):
        return {"error": f"You don't have permission to add {cfg['label']}."}

    if kind == "sales":  # date-keyed, not name-keyed → dedicated path
        return await _commit_sales(db, user, rows)

    created: list[str] = []
    skipped: list[str] = []
    for row in rows:
        name = (row.get("name") or "").strip()
        if not name:
            continue
        fields: dict[str, Any] = {}
        for f in cfg["str_fields"]:
            val = row.get(f)
            if isinstance(val, str) and val.strip():
                fields[f] = val.strip()
        for f in cfg["num_fields"]:
            if row.get(f) is not None:
                d = _dec(row[f])
                if d is not None:
                    fields[f] = d
        # opening stock is worth nothing on the books without a cost — seed the
        # weighted-average so valuation is right from day one.
        if kind == "items" and fields.get("cost_price"):
            fields["average_cost"] = fields["cost_price"]
        try:
            if kind == "items":
                await inventory_service.create_item(db, user.hotel_id, **fields)
            elif kind == "vendors":
                await vendor_service.create_vendor(db, user.hotel_id, **fields)
            elif kind == "recipes":
                await recipe_service.create_recipe(db, user.hotel_id, **fields)
            elif kind == "employees":
                ef = {k: v for k, v in fields.items() if k != "name"}
                ef["full_name"] = name
                await employee_service.create_employee(db, user.hotel_id, **ef)
            else:
                continue
            created.append(name)
        except Exception:  # noqa: BLE001 — duplicate/validation → skip, keep going
            skipped.append(name)

    if created:
        await audit.record(
            db, hotel_id=user.hotel_id, user=user, action=f"assistant.onboard.{kind}",
            summary=f"Copilot onboarding added {len(created)} {cfg['label']}",
            entity_type=kind,
        )
    return {"kind": kind, "created": created, "skipped": skipped}
