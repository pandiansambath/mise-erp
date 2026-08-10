"""Excel export + import for the weekly rota.

Export: the week's shifts as a styled workbook. Template: an empty sheet with the
right headers + an example + the valid employee names. Import: parse a filled
template back into shift dicts (best-effort; never raises on bad cells)."""
import csv
import io
import re
from datetime import date as date_type
from datetime import datetime, time, timedelta
from decimal import Decimal

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.core.xlsx_style import style_table

# DineAI emerald brand for the PDF (RGB).
_BRAND = (5, 150, 105)    # emerald-600
_DARK = (6, 95, 70)       # emerald-900
_MUTED = (100, 116, 139)
_ZEBRA = (236, 253, 245)  # emerald-50
_TOTAL = (209, 250, 229)  # emerald-100

# Excel totals-row look (matches app/core/xlsx_style palette).
_X_TOTAL_BG = "D1FAE5"
_X_TITLE = "065F46"
_x_side = Side(style="thin", color="A7F3D0")
_X_BOX = Border(left=_x_side, right=_x_side, top=_x_side, bottom=_x_side)


def _ps(value) -> str:
    """latin-1 safe text for fpdf2 core fonts."""
    return str(value).encode("latin-1", "replace").decode("latin-1")


def _days(date_from: date_type, date_to: date_type) -> list[date_type]:
    """Every day in the range (capped at 14 so a stray range can't explode the grid)."""
    out: list[date_type] = []
    d = date_from
    while d <= date_to and len(out) < 14:
        out.append(d)
        d += timedelta(days=1)
    return out or [date_from]


def _pivot(shifts: list[dict]) -> dict:
    """Group shifts into an attendance-style grid: one entry per employee, with each
    day's shift time(s) and that employee's total hours."""
    emps: dict = {}
    for s in shifts:
        eid = s["employee_id"]
        e = emps.setdefault(
            eid, {"name": s["employee_name"], "cells": {}, "day_h": {}, "total": Decimal("0")}
        )
        d = s["date"]
        iso = d.isoformat() if hasattr(d, "isoformat") else str(d)
        brk = s.get("break_minutes") or 0
        txt = f'{s["start_time"].strftime("%H:%M")}-{s["end_time"].strftime("%H:%M")}'
        if brk:
            txt += f" -{brk}m"  # unpaid break (already excluded from the hours total)
        e["cells"][iso] = f'{e["cells"][iso]} / {txt}' if iso in e["cells"] else txt
        # The PDF lays the times out itself, so it needs them apart rather than
        # already glued into a string. "09:00-17:00 -30m" is about 30mm of text
        # in a 22mm column, and fpdf's cell() does not wrap — it just runs over
        # the next column. That overlap IS the clumsiness he kept reporting.
        e.setdefault("shifts", {}).setdefault(iso, []).append(
            (
                s["start_time"].strftime("%H:%M"),
                s["end_time"].strftime("%H:%M"),
                int(brk),
            )
        )
        e["day_h"][iso] = e["day_h"].get(iso, Decimal("0")) + s["hours"]
        e["total"] += s["hours"]
    return emps


def _fmt_h(value: Decimal) -> str:
    """Tidy hours: 3.00 -> '3', 3.50 -> '3.5', 0 -> '0'."""
    s = f"{value:.2f}".rstrip("0").rstrip(".")
    return s or "0"


def rota_to_pdf(
    shifts: list[dict], hotel_name: str, date_from, date_to, emp_info: dict | None = None
) -> bytes:
    """A clean, branded weekly rota in the attendance-sheet grid: employees down the
    rows, days across the columns. Landscape so the whole week fits comfortably."""
    emp_info = emp_info or {}
    days = _days(date_from, date_to)
    emps = _pivot(shifts)
    ordered = sorted(emps.items(), key=lambda kv: kv[1]["name"].lower())

    pdf = FPDF(orientation="L")
    # Page breaks are decided row by row below. Left on, fpdf also breaks
    # when the footer is written past its own bottom margin, which is where
    # the stray blank last page came from.
    pdf.set_auto_page_break(False)
    pdf.add_page()
    pw = pdf.w
    m = 12
    # Brand header band
    pdf.set_fill_color(*_BRAND)
    pdf.rect(0, 0, pw, 26, style="F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(m, 6)
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 8, text=_ps(hotel_name), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_x(m)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(
        0, 5, text=_ps(f"WEEKLY ROTA   |   {date_from}  to  {date_to}"),
        new_x=XPos.LMARGIN, new_y=YPos.NEXT,
    )

    # Column geometry — names/role on the left, a column per day, total on the
    # right. The day columns are what everything else gives way to: a rota is
    # read by looking DOWN a day, so the day gets the room.
    avail = pw - 2 * m
    name_w, id_w, total_w = 50, 15, 19
    n_day = len(days)
    day_w = max(16.0, (avail - name_w - id_w - total_w) / n_day)

    rh = 13.0          # tall enough for two stacked times, or two split shifts
    head_h = 10.0
    bottom = pdf.h - 16

    def header_row(y: float) -> float:
        pdf.set_xy(m, y)
        pdf.set_fill_color(*_DARK)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(name_w, head_h, text="  Employee", fill=True)
        pdf.cell(id_w, head_h, text="ID", align="C", fill=True)
        for d in days:
            pdf.cell(day_w, head_h, text=_ps(d.strftime("%a %d/%m")), align="C", fill=True)
        pdf.cell(
            total_w, head_h, text="Hours", align="C", fill=True,
            new_x=XPos.LMARGIN, new_y=YPos.NEXT,
        )
        return y + head_h

    def day_cell(x: float, y: float, entries: list, zebra: bool) -> None:
        """One day for one person.

        The times are STACKED — start above end — instead of written across.
        "09:00" is a third of the width of "09:00-17:00 -30m", so the column
        stops overflowing and the grid can be read down a day at a glance. A
        worked day gets a tinted block so the shape of the week is visible
        before you read a single number."""
        if not entries:
            pdf.set_text_color(210, 218, 226)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_xy(x, y + rh / 2 - 2.5)
            pdf.cell(day_w, 5, text="-", align="C")
            return

        pdf.set_fill_color(*(_ZEBRA if not zebra else (220, 250, 238)))
        pdf.rect(x + 1.1, y + 1.1, day_w - 2.2, rh - 2.2, style="F", round_corners=True,
                 corner_radius=1.6)

        split = len(entries) > 1
        size = 6.2 if split else 8.4
        line = 3.1 if split else 4.3
        top = y + (rh - line * 2 * len(entries)) / 2

        for k, (st, en, brk) in enumerate(entries[:2]):
            pdf.set_text_color(*_DARK)
            pdf.set_font("Helvetica", "B", size)
            pdf.set_xy(x, top + k * line * 2)
            pdf.cell(day_w, line, text=_ps(st), align="C")
            pdf.set_font("Helvetica", "", size)
            pdf.set_text_color(*_MUTED)
            pdf.set_xy(x, top + k * line * 2 + line)
            pdf.cell(day_w, line, text=_ps(en), align="C")
            # An unpaid break is marked with a dot in the corner of the
            # block. Appending it to the time ("17:30.") read as a full
            # stop, and spelling out "-30m" is what overflowed the column
            # in the first place.
            if brk:
                pdf.set_fill_color(129, 199, 174)
                pdf.circle(x + day_w - 4.4, y + 3.0, 0.75, style="F")

        if split:
            pdf.set_draw_color(168, 214, 195)
            pdf.set_line_width(0.15)
            mid = top + line * 2
            pdf.line(x + day_w * 0.28, mid, x + day_w * 0.72, mid)

        if len(entries) > 2:
            pdf.set_font("Helvetica", "B", 5.4)
            pdf.set_text_color(*_MUTED)
            pdf.set_xy(x, y + rh - 3.4)
            pdf.cell(day_w, 3, text=_ps(f"+{len(entries) - 2}"), align="C")

    y = header_row(32.0)

    daily = {d.isoformat(): Decimal("0") for d in days}
    grand = Decimal("0")
    has_break = False

    if not ordered:
        pdf.set_xy(m, y)
        pdf.set_text_color(*_MUTED)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, rh, text="  No shifts scheduled for this week.")
        y += rh

    for i, (eid, e) in enumerate(ordered):
        # Page breaks were never handled: a long team simply ran off the bottom
        # of the sheet. Carry the header over so page two is still readable.
        if y + rh > bottom:
            pdf.add_page()
            y = header_row(20.0)

        code, title = emp_info.get(eid, ("", ""))
        zebra = i % 2 == 1
        if zebra:
            pdf.set_fill_color(248, 252, 250)
            pdf.rect(m, y, avail, rh, style="F")

        pdf.set_text_color(*_DARK)
        pdf.set_font("Helvetica", "B", 8.6)
        pdf.set_xy(m, y + rh / 2 - 4.6)
        pdf.cell(name_w, 5, text=f"  {_ps(e['name'])}")
        pdf.set_font("Helvetica", "", 6.6)
        pdf.set_text_color(*_MUTED)
        pdf.set_xy(m, y + rh / 2 + 0.2)
        pdf.cell(name_w, 4, text=f"  {_ps(title or '-')}")

        pdf.set_font("Helvetica", "", 7.4)
        pdf.set_xy(m + name_w, y + rh / 2 - 2.2)
        pdf.cell(id_w, 5, text=_ps(code or "-"), align="C")

        x = m + name_w + id_w
        for d in days:
            iso = d.isoformat()
            entries = (e.get("shifts") or {}).get(iso, [])
            if any(b for _, _, b in entries):
                has_break = True
            day_cell(x, y, entries, zebra)
            daily[iso] += e["day_h"].get(iso, Decimal("0"))
            x += day_w

        grand += e["total"]
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*_BRAND)
        pdf.set_xy(x, y + rh / 2 - 2.4)
        pdf.cell(total_w, 5, text=_ps(_fmt_h(e["total"]) + "h"), align="C")

        pdf.set_draw_color(226, 240, 234)
        pdf.set_line_width(0.2)
        pdf.line(m, y + rh, m + avail, y + rh)
        y += rh

    if ordered:
        if y + rh > bottom:
            pdf.add_page()
            y = header_row(20.0)
        pdf.set_fill_color(*_TOTAL)
        pdf.rect(m, y, avail, rh - 2, style="F", round_corners=True, corner_radius=1.6)
        pdf.set_text_color(*_DARK)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_xy(m, y + 2)
        pdf.cell(name_w + id_w, 6, text="  Hours per day")
        x = m + name_w + id_w
        for d in days:
            pdf.set_xy(x, y + 2)
            pdf.cell(day_w, 6, text=_ps(_fmt_h(daily[d.isoformat()])), align="C")
            x += day_w
        pdf.set_xy(x, y + 2)
        pdf.cell(total_w, 6, text=_ps(_fmt_h(grand) + "h"), align="C")
        y += rh

    # The legend earns its line only when the sheet actually uses the marker.
    pdf.set_text_color(*_MUTED)
    pdf.set_font("Helvetica", "", 6.8)
    if has_break:
        pdf.set_xy(m, y + 1)
        pdf.cell(
            0, 4,
            text="A dot in the corner of a shift means an unpaid break is already "
                 "taken out of the hours.",
        )

    pdf.set_y(pdf.h - 12)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5, text="Generated by DineAI - every plate, every penny", align="C")
    return bytes(pdf.output())


def rota_to_xlsx(
    shifts: list[dict], date_from: date_type, date_to: date_type, emp_info: dict | None = None
) -> bytes:
    """Weekly rota as the attendance-sheet grid: employees down rows, days across
    columns, each employee's total hours, and a daily-totals row at the bottom."""
    emp_info = emp_info or {}
    days = _days(date_from, date_to)
    emps = _pivot(shifts)
    ordered = sorted(emps.items(), key=lambda kv: kv[1]["name"].lower())
    n_day = len(days)
    total_col = 3 + n_day + 1  # 1-based index of the "Total h" column

    wb = Workbook()
    ws = wb.active
    ws.title = "Rota"
    day_headers = [d.strftime("%a %d/%m") for d in days]
    headers = ["Employee", "Emp ID", "Role", *day_headers, "Total h"]

    daily = {d.isoformat(): Decimal("0") for d in days}
    grand = Decimal("0")
    r = 4
    for eid, e in ordered:
        code, title = emp_info.get(eid, ("", ""))
        ws.cell(row=r, column=1, value=e["name"])
        ws.cell(row=r, column=2, value=code or "")
        ws.cell(row=r, column=3, value=title or "")
        for j, d in enumerate(days):
            iso = d.isoformat()
            ws.cell(row=r, column=4 + j, value=e["cells"].get(iso, "—"))
            daily[iso] += e["day_h"].get(iso, Decimal("0"))
        ws.cell(row=r, column=total_col, value=float(e["total"]))
        grand += e["total"]
        r += 1

    style_table(
        ws, title="DineAI — Weekly Rota",
        subtitle=f"{date_from} → {date_to}   ·   '-30m' after a shift = 30-min unpaid break",
        headers=headers, n_rows=len(ordered),
        widths=[22, 10, 16, *([17] * n_day), 10], right_cols={total_col},
    )

    if not ordered:
        note = ws.cell(row=4, column=1, value="No shifts scheduled for this week.")
        note.font = Font(italic=True, color=_X_TITLE)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Center the day columns + Emp ID/Role (header + data) for a clean grid.
    for rr in range(3, 4 + len(ordered)):
        for col in range(4, 4 + n_day):
            ws.cell(row=rr, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        ws.cell(row=rr, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=rr, column=3).alignment = Alignment(horizontal="center")

    # Daily-totals row (bold emerald on a light fill), matching the styler.
    tr = 4 + len(ordered)
    bold = Font(bold=True, color=_X_TITLE)
    fill = PatternFill("solid", fgColor=_X_TOTAL_BG)
    label = ws.cell(row=tr, column=1, value="Daily total (hours)")
    label.font, label.fill, label.border = bold, fill, _X_BOX
    for col in (2, 3):
        c = ws.cell(row=tr, column=col)
        c.fill, c.border = fill, _X_BOX
    for j, d in enumerate(days):
        c = ws.cell(row=tr, column=4 + j, value=float(daily[d.isoformat()]))
        c.font, c.fill, c.border = bold, fill, _X_BOX
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=tr, column=total_col, value=float(grand))
    c.font, c.fill, c.border = bold, fill, _X_BOX
    c.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _grid_headers(days: list[date_type]) -> list[str]:
    return ["Employee", "Emp ID", "Role", *(d.strftime("%a %d/%m") for d in days), "Total h"]


def _grid_subtitle(date_from: date_type, date_to: date_type) -> str:
    # NOTE: the import parser reads the ISO date range from the START of this line.
    return (
        f"{date_from} → {date_to}   ·   fill each cell like 09:00-17:00 "
        "(add -30m for a 30-min unpaid break); leave blank for a day off, then upload"
    )


def template_grid_xlsx(
    emp_rows: list[tuple[str, str, str]], date_from: date_type, date_to: date_type
) -> bytes:
    """A blank weekly-rota GRID that looks exactly like the download — employees down
    the rows, days across the columns. The owner fills the cells and re-uploads it."""
    days = _days(date_from, date_to)
    n_day = len(days)
    total_col = 3 + n_day + 1
    rows = emp_rows or [("Staff full name", "", "")]
    wb = Workbook()
    ws = wb.active
    ws.title = "Rota"
    for i, (name, code, title) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=code or "")
        ws.cell(row=r, column=3, value=title or "")
    style_table(
        ws, title="DineAI — Weekly Rota", subtitle=_grid_subtitle(date_from, date_to),
        headers=_grid_headers(days), n_rows=len(rows),
        widths=[22, 10, 16, *([17] * n_day), 10], right_cols={total_col},
    )
    for rr in range(3, 4 + len(rows)):
        for col in range(4, 4 + n_day):
            ws.cell(row=rr, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        ws.cell(row=rr, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=rr, column=3).alignment = Alignment(horizontal="center")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_grid_csv(
    emp_rows: list[tuple[str, str, str]], date_from: date_type, date_to: date_type
) -> bytes:
    """Same grid as a CSV (title + instruction + header, one row per employee)."""
    days = _days(date_from, date_to)
    rows = emp_rows or [("Staff full name", "", "")]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["DineAI — Weekly Rota"])
    w.writerow([_grid_subtitle(date_from, date_to)])
    w.writerow(_grid_headers(days))
    for name, code, title in rows:
        w.writerow([name, code or "", title or "", *([""] * len(days)), ""])
    return buf.getvalue().encode("utf-8-sig")


def _parse_time(v) -> time | None:
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, time):
        return v
    s = str(v or "").strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p", "%I%p"):
        try:
            return datetime.strptime(s.upper(), fmt).time()
        except ValueError:
            continue
    return None


_DAY_HDR = re.compile(r"^[A-Za-z]{3}\s+(\d{1,2})/(\d{1,2})$")
_ISO_DATE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")
# A shift cell: '09:00-17:00' with an optional ' -30m' unpaid-break suffix.
_CELL = re.compile(r"^\s*(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})(?:\s*[-–]?\s*(\d+)\s*m)?\s*$")
_EMPTY_CELL = {"", "—", "-", "–", "off", "day off", "—/—"}


def _read_rows(data: bytes, filename: str, content_type: str) -> list[list]:
    """Best-effort 2-D rows from an xlsx or csv upload (never raises)."""
    name = (filename or "").lower()
    is_csv = name.endswith(".csv") or "csv" in (content_type or "").lower()
    try:
        if is_csv:
            text = data.decode("utf-8-sig", errors="replace")
            return [list(r) for r in csv.reader(io.StringIO(text))]
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return [list(r) for r in wb.active.iter_rows(values_only=True)]
    except Exception:  # noqa: BLE001 — unreadable → empty
        return []


def parse_rota_grid(
    data: bytes, filename: str, content_type: str
) -> tuple[dict | None, list[str]]:
    """Read a filled weekly-rota GRID (the download/template layout) back into shifts.

    Returns (grid, errors) where grid = {"shifts", "employees", "from", "to"}, or
    (None, []) when the file is NOT a grid (so the caller can fall back to the older
    one-row-per-shift template). Cells look like '09:00-17:00' or '09:00-17:00 -30m';
    a day can hold two shifts split by ' / '. Blank or '—' means a day off."""
    rows = _read_rows(data, filename, content_type)
    if not rows:
        return None, []
    hr = next(
        (i for i, row in enumerate(rows[:8])
         if row and any(str(c or "").strip().lower() == "employee" for c in row)),
        None,
    )
    if hr is None:
        return None, []
    header = [str(c or "").strip() for c in rows[hr]]
    lower = [h.lower() for h in header]
    if "date" in lower and "start" in lower:
        return None, []  # the older one-row-per-shift template → let caller fall back
    emp_col = lower.index("employee")
    # Year/month come from the ISO date range printed in the subtitle line above.
    head_text = " ".join(str(c or "") for r in rows[: hr + 1] for c in r)
    iso = _ISO_DATE.search(head_text)
    base_year = int(iso.group(1)) if iso else date_type.today().year
    base_month = int(iso.group(2)) if iso else None
    day_cols: list[tuple[int, date_type]] = []
    for ci, h in enumerate(header):
        dm = _DAY_HDR.match(h)
        if not dm:
            continue
        dd, mm = int(dm.group(1)), int(dm.group(2))
        yr = base_year + 1 if (base_month == 12 and mm == 1) else base_year
        try:
            day_cols.append((ci, date_type(yr, mm, dd)))
        except ValueError:
            continue
    if not day_cols:
        return None, []
    shifts: list[dict] = []
    employees: list[str] = []
    errors: list[str] = []
    for row in rows[hr + 1:]:
        if not row or emp_col >= len(row):
            continue
        name = str(row[emp_col] or "").strip()
        if not name:
            continue
        if name.lower().startswith("daily total"):
            break
        employees.append(name)
        for ci, d in day_cols:
            if ci >= len(row):
                continue
            text = str(row[ci] or "").strip()
            if text.lower() in _EMPTY_CELL:
                continue
            for part in (p.strip() for p in text.split("/") if p.strip()):
                cm = _CELL.match(part)
                st = _parse_time(cm.group(1)) if cm else None
                en = _parse_time(cm.group(2)) if cm else None
                if not cm or not st or not en:
                    errors.append(
                        f"{name} · {d.strftime('%a %d/%m')}: couldn't read “{part}” "
                        "(use 09:00-17:00, add -30m for a break)"
                    )
                    continue
                shifts.append({
                    "employee": name, "date": d, "start": st, "end": en,
                    "break_minutes": int(cm.group(3) or 0),
                })
    return (
        {
            "shifts": shifts,
            "employees": employees,
            "from": min(d for _, d in day_cols),
            "to": max(d for _, d in day_cols),
        },
        errors,
    )
