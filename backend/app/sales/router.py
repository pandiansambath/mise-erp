"""Daily sales & cash endpoints. Hotel-scoped."""
import io
import re
import uuid
from datetime import date as date_type
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.deps import require
from app.auth.models import User
from app.core.config import settings
from app.core.database import get_db
from app.hotels.models import Hotel
from app.sales import pdf as sales_pdf
from app.sales import service
from app.sales.schemas import (
    ChannelCreate,
    ChannelOut,
    ChannelUpdate,
    DayCreatedOut,
    DaySummary,
    DayUpsert,
    DishCount,
    DishSalesIn,
    DishSalesOut,
    LineCreate,
    RangeSummary,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

router = APIRouter(prefix="/sales", tags=["sales"])


# ── Channels ────────────────────────────────────────────────────────────────
@router.get("/channels", response_model=list[ChannelOut])
async def list_channels(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:read")),
) -> list[ChannelOut]:
    await service.ensure_default_channels(db, user.hotel_id)  # idempotent
    channels = await service.list_channels(db, user.hotel_id, active_only=False)
    counts = await service.channel_usage_counts(db, user.hotel_id)
    return [
        ChannelOut(
            id=c.id, name=c.name, commission_pct=c.commission_pct, is_active=c.is_active,
            usage_count=counts.get(c.id, 0),
        )
        for c in channels
    ]


@router.post("/channels", response_model=ChannelOut, status_code=status.HTTP_201_CREATED)
async def create_channel(
    payload: ChannelCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:config")),
) -> ChannelOut:
    ch = await service.create_channel(db, user.hotel_id, payload.name, payload.commission_pct)
    return ChannelOut.model_validate(ch)


@router.patch("/channels/{channel_id}", response_model=ChannelOut)
async def update_channel(
    channel_id: uuid.UUID,
    payload: ChannelUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:config")),
) -> ChannelOut:
    ch = await service.get_channel(db, channel_id, user.hotel_id)
    if ch is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Channel not found")
    ch = await service.update_channel(db, ch, **payload.model_dump(exclude_unset=True))
    return ChannelOut.model_validate(ch)


# ── Excel import (daily sales) ───────────────────────────────────────────────
def _parse_sales_rows(data: bytes) -> list[tuple[str, Decimal | None, str]]:
    """Read an .xlsx with columns Channel / Gross (+ optional Method=CASH|CARD)."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any parse failure -> friendly 400
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "Could not read the file — upload a .xlsx Excel file."
        ) from exc
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return []
    cols = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}

    def find(*names: str) -> int | None:
        return next((cols[n] for n in names if n in cols), None)

    ci = find("channel", "source", "platform")
    gi = find("gross", "gross amount", "amount", "sales", "total")
    mi = find("method", "payment", "payment method", "type")
    if ci is None or gi is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "Excel needs a 'Channel' column and a 'Gross' column."
        )
    out: list[tuple[str, Decimal | None, str]] = []
    for row in rows_iter:
        ch = row[ci] if ci < len(row) else None
        raw = row[gi] if gi < len(row) else None
        m = row[mi] if (mi is not None and mi < len(row)) else None
        if ch is None and raw is None:
            continue
        gross: Decimal | None = None
        if raw is not None:
            cleaned = re.sub(r"[^0-9.]", "", str(raw))
            try:
                gross = Decimal(cleaned) if cleaned else None
            except InvalidOperation:
                gross = None
        method = "CASH" if (m is not None and "cash" in str(m).lower()) else "CARD"
        out.append((str(ch).strip() if ch is not None else "", gross, method))
    return out


@router.get("/sales-template.xlsx")
async def sales_template(user: User = Depends(require("sales:read"))) -> Response:
    """A sample Excel for the daily-sales import (Channel, Gross, Method)."""
    from openpyxl import Workbook

    from app.core.xlsx_style import style_table

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    rows = [
        ["Dine-in", 240.00, "CARD"], ["Deliveroo", 86.50, "CARD"], ["Cash counter", 120.00, "CASH"],
    ]
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            ws.cell(row=4 + i, column=c, value=v)
    style_table(
        ws, title="Mise — Sales import template", headers=["Channel", "Gross", "Method"],
        n_rows=len(rows), subtitle="One row per channel for a day, then upload on Sales & Cash",
        widths=[22, 14, 12], right_cols={2},
    )
    buf = io.BytesIO()
    wb.save(buf)
    fname = "mise-sales-template.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/days/{day}/import")
async def import_day_sales(
    day: date_type,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:write")),
) -> dict:
    """Upload a day's sales as Excel. Channels matched by name; unknown skipped."""
    data = await file.read()
    if len(data) > settings.max_upload_mb * 1024 * 1024:
        raise HTTPException(
            status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, f"File exceeds {settings.max_upload_mb} MB"
        )
    rows = _parse_sales_rows(data)
    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No rows found in the file.")
    record = await service.upsert_day(db, user.hotel_id, day, entered_by=user.id)
    added = 0
    skipped: list[str] = []
    for name, gross, method in rows:
        if not name or gross is None or gross < 0:
            if name:
                skipped.append(name)
            continue
        ch = await service.get_channel_by_name(db, user.hotel_id, name)
        if ch is None:
            skipped.append(f"{name} (unknown channel)")
            continue
        await service.add_line(db, record, ch.id, gross, method)
        added += 1
    return {"added": added, "skipped": skipped}


# ── Daily entry ───────────────────────────────────────────────────────────────
@router.get("/days/{day}", response_model=DaySummary)
async def get_day(
    day: date_type,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:read")),
) -> DaySummary:
    return DaySummary.model_validate(await service.day_summary(db, user.hotel_id, day))


@router.get("/days/{day}/sheet.pdf")
async def day_sheet_pdf(
    day: date_type,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:read")),
) -> Response:
    summary = await service.day_summary(db, user.hotel_id, day)
    hotel = await db.get(Hotel, user.hotel_id)
    pdf = sales_pdf.generate_day_pdf(summary, hotel)
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="sales-{day}.pdf"'},
    )


@router.patch("/days/{day}", response_model=DayCreatedOut)
async def upsert_day(
    day: date_type,
    payload: DayUpsert,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:write")),
) -> DayCreatedOut:
    record = await service.upsert_day(
        db,
        user.hotel_id,
        day,
        opening_cash=payload.opening_cash,
        cash_counted=payload.cash_counted,
        notes=payload.notes,
        entered_by=user.id,
    )
    return DayCreatedOut.model_validate(record)


@router.post("/days/{day}/lines", response_model=DaySummary, status_code=status.HTTP_201_CREATED)
async def add_line(
    day: date_type,
    payload: LineCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:write")),
) -> DaySummary:
    channel = await service.get_channel(db, payload.channel_id, user.hotel_id)
    if channel is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Channel not found")
    record = await service.upsert_day(db, user.hotel_id, day, entered_by=user.id)
    await service.add_line(
        db, record, payload.channel_id, payload.gross_amount, payload.payment_method, payload.notes
    )
    return DaySummary.model_validate(await service.day_summary(db, user.hotel_id, day))


@router.delete("/days/{day}/lines/{line_id}", response_model=DaySummary)
async def delete_line(
    day: date_type,
    line_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:write")),
) -> DaySummary:
    line = await service.get_line(db, line_id)
    record = await service._get_day(db, user.hotel_id, day)
    if line is None or record is None or line.daily_sales_id != record.id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Line not found")
    await service.delete_line(db, line)
    return DaySummary.model_validate(await service.day_summary(db, user.hotel_id, day))


@router.get("/summary", response_model=RangeSummary)
async def summary(
    date_from: date_type = Query(...),
    date_to: date_type = Query(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:read")),
) -> RangeSummary:
    return RangeSummary.model_validate(
        await service.range_summary(db, user.hotel_id, date_from, date_to)
    )


# ── Dish sales (menu-engineering bridge) ──────────────────────────────────────
@router.get("/dishes/{day}", response_model=DishSalesOut)
async def get_dish_sales(
    day: date_type,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:read")),
) -> DishSalesOut:
    counts = await service.list_dish_sales(db, user.hotel_id, day)
    return DishSalesOut(date=day, counts=[DishCount(recipe_id=k, qty=v) for k, v in counts.items()])


@router.post("/dishes/{day}", response_model=DishSalesOut)
async def set_dish_sales(
    day: date_type,
    payload: DishSalesIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require("sales:write")),
) -> DishSalesOut:
    await service.upsert_dish_sales(
        db, user.hotel_id, day, {c.recipe_id: c.qty for c in payload.counts}
    )
    counts = await service.list_dish_sales(db, user.hotel_id, day)
    return DishSalesOut(date=day, counts=[DishCount(recipe_id=k, qty=v) for k, v in counts.items()])
