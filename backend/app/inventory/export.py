"""CSV / Excel export for inventory: stock-on-hand valuation and the waste log.
Amounts are in the hotel's base currency (the display-currency toggle is view-only)."""
import csv
import io
from datetime import date as date_type
from decimal import Decimal

from openpyxl import Workbook

from app.core.xlsx_style import style_table, total_row

_Q2 = Decimal("0.01")


def _value(item) -> Decimal:
    return (item.current_stock * item.average_cost).quantize(_Q2)


def _exact_price(item, suppliers: dict | None):
    """The CHOSEN supplier's current £/unit — the exact price you pay now (vs the
    blended weighted-average cost). None when no supplier is chosen for the item."""
    chosen = (suppliers or {}).get(item.id)
    return chosen[2] if chosen and len(chosen) > 2 and chosen[2] is not None else None


def _exact_value(item, suppliers: dict | None):
    """Stock-on-hand valued at the exact (chosen-supplier) price."""
    p = _exact_price(item, suppliers)
    return (item.current_stock * p).quantize(_Q2) if p is not None else None


def _supplier(item, suppliers: dict | None) -> str:
    chosen = (suppliers or {}).get(item.id)
    if not chosen:
        return ""
    name, is_chosen = chosen[0], chosen[1]
    return f"{'★ ' if is_chosen else ''}{name}"


# ── Inventory stock valuation ────────────────────────────────────────────────
# Two costings, side by side, each clearly labelled:
#   Avg cost            = blended weighted-average of stock ON HAND → its real value now.
#   Current buy price   = chosen supplier's £/unit TODAY → what it costs to buy more.
# (Per-batch historical prices live in each item's purchase history, not here.)
_ITEM_COLS = [
    "Item", "Category", "In stock", "Unit", "Min stock",
    "Avg cost", "Stock value (avg)", "Current buy price", "Value at current price",
    "Supplier", "Status",
]


def items_to_csv(items, suppliers: dict | None = None) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Mise — Inventory stock valuation"])
    w.writerow([])
    w.writerow(_ITEM_COLS)
    total = Decimal("0")
    total_exact = Decimal("0")
    for it in items:
        val = _value(it)
        total += val
        ep = _exact_price(it, suppliers)
        ev = _exact_value(it, suppliers)
        if ev is not None:
            total_exact += ev
        w.writerow([
            it.name, it.category or "", str(it.current_stock), it.unit,
            str(it.min_stock_level) if it.min_stock_level is not None else "",
            str(it.average_cost), str(val),
            str(ep) if ep is not None else "", str(ev) if ev is not None else "",
            _supplier(it, suppliers), "active" if it.is_active else "archived",
        ])
    w.writerow([])
    w.writerow([
        "", "", "", "", "", "Total stock value", str(total.quantize(_Q2)),
        "", str(total_exact.quantize(_Q2)),
    ])
    return buf.getvalue().encode("utf-8")


def items_to_xlsx(items, suppliers: dict | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock valuation"
    r = 4
    total = 0.0
    total_exact = 0.0
    for it in items:
        val = float(_value(it))
        total += val
        ep = _exact_price(it, suppliers)
        ev = _exact_value(it, suppliers)
        if ev is not None:
            total_exact += float(ev)
        ws.cell(row=r, column=1, value=it.name)
        ws.cell(row=r, column=2, value=it.category or "")
        ws.cell(row=r, column=3, value=float(it.current_stock))
        ws.cell(row=r, column=4, value=it.unit)
        min_v = float(it.min_stock_level) if it.min_stock_level is not None else None
        ws.cell(row=r, column=5, value=min_v)
        ws.cell(row=r, column=6, value=float(it.average_cost)).number_format = "#,##0.0000"
        ws.cell(row=r, column=7, value=val).number_format = "#,##0.00"
        ws.cell(
            row=r, column=8, value=float(ep) if ep is not None else None
        ).number_format = "#,##0.0000"
        ws.cell(
            row=r, column=9, value=float(ev) if ev is not None else None
        ).number_format = "#,##0.00"
        ws.cell(row=r, column=10, value=_supplier(it, suppliers))
        ws.cell(row=r, column=11, value="active" if it.is_active else "archived")
        r += 1
    style_table(
        ws, title="Mise — Inventory stock valuation", headers=_ITEM_COLS, n_rows=len(items),
        widths=[26, 16, 10, 8, 10, 12, 15, 16, 22, 18, 10], right_cols={3, 5, 6, 7, 8, 9},
    )
    total_row(
        ws, row=r + 1, label="Total stock value", label_col=6,
        value=round(total, 2), value_col=7,
    )
    total_row(ws, row=r + 1, label="", label_col=8, value=round(total_exact, 2), value_col=9)
    ws.cell(row=r + 1, column=7).number_format = "#,##0.00"
    ws.cell(row=r + 1, column=9).number_format = "#,##0.00"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ── Waste log ────────────────────────────────────────────────────────────────
_WASTE_COLS = ["Date", "Item", "Qty", "Unit", "Unit cost", "Value", "Reason"]


def _title(date_from: date_type | None, date_to: date_type | None) -> str:
    if date_from and date_to:
        return f"Mise — Waste log ({date_from} to {date_to})"
    return "Mise — Waste log"


def _wdate(created) -> str:
    return created.date().isoformat() if hasattr(created, "date") else str(created)[:10]


def waste_to_csv(rows: list[dict], date_from=None, date_to=None) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([_title(date_from, date_to)])
    w.writerow([])
    w.writerow(_WASTE_COLS)
    total = Decimal("0")
    for r in rows:
        total += Decimal(str(r["value"]))
        w.writerow([
            _wdate(r["created_at"]), r["item_name"], str(r["quantity"]), r["unit"],
            str(r["unit_cost"]) if r["unit_cost"] is not None else "", str(r["value"]),
            r["reason"] or "",
        ])
    w.writerow([])
    w.writerow(["", "", "", "", "Total waste", str(total.quantize(_Q2))])
    return buf.getvalue().encode("utf-8")


def waste_to_xlsx(rows: list[dict], date_from=None, date_to=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Waste log"
    r = 4
    total = 0.0
    for row in rows:
        total += float(row["value"])
        ws.cell(row=r, column=1, value=_wdate(row["created_at"]))
        ws.cell(row=r, column=2, value=row["item_name"])
        ws.cell(row=r, column=3, value=float(row["quantity"]))
        ws.cell(row=r, column=4, value=row["unit"])
        uc = float(row["unit_cost"]) if row["unit_cost"] is not None else None
        ws.cell(row=r, column=5, value=uc).number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=float(row["value"])).number_format = "#,##0.00"
        ws.cell(row=r, column=7, value=row["reason"] or "")
        r += 1
    style_table(
        ws, title=_title(date_from, date_to), headers=_WASTE_COLS, n_rows=len(rows),
        widths=[12, 26, 10, 10, 12, 14, 22], right_cols={3, 5, 6},
    )
    total_row(ws, row=r + 1, label="Total waste", label_col=5, value=round(total, 2), value_col=6)
    ws.cell(row=r + 1, column=6).number_format = "#,##0.00"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
