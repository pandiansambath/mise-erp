"""Strict import templates, shared across the app.

Generate a fill-in template (.xlsx or .csv) from a column spec, and validate an
uploaded file against that exact spec — returning precise, human-friendly errors
("Row 5: 'Cost price' must be a number") so the user can fix the file and re-upload.

Only tabular formats (.xlsx/.csv) are validated here: they round-trip exactly, so we
can prove a match. Freeform docs (PDF/Word/photo) can't be strictly checked — those
go through the AI reader instead (app.assistant.ingest)."""
from __future__ import annotations

import csv as _csv
import io
from dataclasses import dataclass, field
from datetime import date as _date
from datetime import datetime as _datetime
from datetime import time as _time
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from app.core.xlsx_style import style_table

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_ROWS = 1000


@dataclass
class Column:
    key: str                      # canonical field name used in the parsed row
    header: str                   # the column title shown in the template
    required: bool = False
    kind: str = "text"            # "text" | "number"
    aliases: tuple[str, ...] = ()  # other accepted header spellings
    example: str = ""


@dataclass
class TemplateSpec:
    name: str                     # e.g. "Inventory items"
    columns: list[Column]
    subtitle: str = ""
    sample_rows: list[list] = field(default_factory=list)

    def headers(self) -> list[str]:
        return [c.header for c in self.columns]

    def required_headers(self) -> list[str]:
        return [c.header for c in self.columns if c.required]


def _norm(v) -> str:
    return str(v if v is not None else "").strip().lower()


def _coerce_date(raw) -> str | None:
    """Parse a date cell (Excel date object or text) → 'YYYY-MM-DD', or None."""
    if isinstance(raw, _datetime):
        return raw.date().isoformat()
    if isinstance(raw, _date):
        return raw.isoformat()
    s = str(raw or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return _datetime.strptime(s[:11].strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _coerce_time(raw) -> str | None:
    """Parse a time cell (Excel time object or text) → 'HH:MM', or None."""
    if isinstance(raw, _datetime):
        return raw.strftime("%H:%M")
    if isinstance(raw, _time):
        return raw.strftime("%H:%M")
    s = str(raw or "").strip().upper()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p", "%I %p", "%I%p"):
        try:
            return _datetime.strptime(s, fmt).strftime("%H:%M")
        except ValueError:
            continue
    return None


# ── Template generation ───────────────────────────────────────────────────────
def template_xlsx(spec: TemplateSpec) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = spec.name[:31]
    for i, row in enumerate(spec.sample_rows):
        for c, val in enumerate(row, start=1):
            ws.cell(row=4 + i, column=c, value=val)
    req = set(spec.required_headers())
    headers = [f"{h} *" if h in req else h for h in spec.headers()]
    right = {i for i, c in enumerate(spec.columns, start=1) if c.kind == "number"}
    style_table(
        ws, title=f"Mise — {spec.name} template",
        subtitle=spec.subtitle or "Fill the rows, keep the headers, then upload. * = required.",
        headers=headers, n_rows=max(len(spec.sample_rows), 1),
        widths=[max(12, len(h) + 4) for h in headers], right_cols=right,
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_csv(spec: TemplateSpec) -> bytes:
    buf = io.StringIO()
    w = _csv.writer(buf)
    req = set(spec.required_headers())
    w.writerow([f"{h} *" if h in req else h for h in spec.headers()])
    for row in spec.sample_rows:
        w.writerow(row)
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens it cleanly


def template_pdf(spec: TemplateSpec) -> bytes:
    """A printable REFERENCE of the template (columns + examples). PDFs can't be
    uploaded back (no fixed grid) — fill the Excel/CSV to import. fpdf2 (prod) only."""
    from fpdf.enums import XPos, YPos

    from app.core.pdf import DARK, ZEBRA, branded_pdf, footer, ps, table_header

    pdf = branded_pdf(f"{spec.name} template", "Reference only — fill the Excel or CSV to import.")
    avail = pdf.w - 28
    cols = spec.columns
    w_each = avail / max(len(cols), 1)
    req = set(spec.required_headers())
    table_header(pdf, [(f"{c.header}{' *' if c.header in req else ''}", w_each, "L") for c in cols])
    pdf.set_font("Helvetica", "", 8)
    for i, row in enumerate(spec.sample_rows):
        pdf.set_x(14)
        pdf.set_fill_color(*ZEBRA)
        for j in range(len(cols)):
            val = row[j] if j < len(row) else ""
            pdf.cell(w_each, 8, text=ps(str(val)), border="B", fill=i % 2 == 1)
        pdf.ln(8)
    pdf.ln(4)
    pdf.set_text_color(*DARK)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_x(14)
    pdf.multi_cell(
        0, 5,
        text=ps("* = required. To import: download the Excel or CSV template, fill one row "
                "per record keeping the headers, and upload it on the page."),
        new_x=XPos.LMARGIN, new_y=YPos.NEXT,
    )
    footer(pdf)
    return bytes(pdf.output())


# ── Validation ────────────────────────────────────────────────────────────────
def _read_rows(file_bytes: bytes, filename: str, mime: str) -> list[list] | None:
    """Return the sheet as a list of rows, or None if it isn't a readable .xlsx/.csv."""
    name = (filename or "").lower()
    is_xlsx = name.endswith(".xlsx") or (mime or "").split(";")[0].strip() == XLSX_MIME
    is_csv = name.endswith(".csv") or (mime or "").startswith("text/")
    try:
        if is_xlsx:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            return [list(r) for r in ws.iter_rows(values_only=True)]
        if is_csv:
            text = file_bytes.decode("utf-8-sig", "replace")
            return [r for r in _csv.reader(io.StringIO(text))]
    except Exception:  # noqa: BLE001 — unreadable → treated as "not the template"
        return None
    return None


def _match_header(spec: TemplateSpec, rows: list[list]) -> tuple[int, dict[str, int]] | None:
    """Find the header row + map each spec column to its column index. Tolerates a
    styled template (header on row 3) and a trailing ' *' on required headers."""
    wanted = {}
    for ci, col in enumerate(spec.columns):
        for h in (col.header, *col.aliases):
            wanted[_norm(h)] = ci
    best = None
    for ri, row in enumerate(rows[:8]):
        found: dict[str, int] = {}
        for idx, cell in enumerate(row):
            key = _norm(cell).removesuffix(" *").strip()
            if key in wanted:
                found[spec.columns[wanted[key]].key] = idx
        if best is None or len(found) > len(best[1]):
            best = (ri, found)
    return best


def parse_upload(
    file_bytes: bytes, filename: str, mime: str, spec: TemplateSpec
) -> tuple[list[dict], list[str]]:
    """Validate an uploaded template. Returns (rows, errors). If errors is non-empty,
    rows is empty — show the errors and let the user fix + re-upload."""
    rows = _read_rows(file_bytes, filename, mime)
    if rows is None:
        return [], [
            "This isn't a readable Excel (.xlsx) or CSV file. Download the template and "
            "fill it in — or use “Import with AI” for a PDF, Word doc or photo."
        ]
    match = _match_header(spec, rows)
    if not match or not match[1]:
        return [], [
            "Couldn't find the template's header row. Use the template unchanged — it "
            "needs these columns: " + ", ".join(spec.headers()) + "."
        ]
    header_idx, colmap = match
    missing = [c.header for c in spec.columns if c.required and c.key not in colmap]
    if missing:
        return [], [
            f"Missing required column{'s' if len(missing) > 1 else ''}: "
            + ", ".join(f"“{m}”" for m in missing)
            + ". The template needs: " + ", ".join(spec.headers()) + "."
        ]

    by_key = {c.key: c for c in spec.columns}
    out: list[dict] = []
    errors: list[str] = []
    data = rows[header_idx + 1:][:MAX_ROWS]
    for n, row in enumerate(data, start=1):
        if not any(_norm(c) for c in row):
            continue  # blank line
        rec: dict = {}
        row_errs: list[str] = []
        for key, idx in colmap.items():
            col = by_key[key]
            raw = row[idx] if idx < len(row) else None
            sval = str(raw).strip() if raw is not None else ""
            if not sval:
                if col.required:
                    row_errs.append(f"“{col.header}” is required")
                continue
            if col.kind == "number":
                cleaned = sval.replace(",", "").lstrip("£$€ ").strip()
                try:
                    rec[key] = float(Decimal(cleaned))
                except (InvalidOperation, ValueError):
                    row_errs.append(f"“{col.header}” must be a number (got “{sval}”)")
            elif col.kind == "date":
                d = _coerce_date(raw)
                if d is None:
                    row_errs.append(f"“{col.header}” must be a date like 2026-06-30 (got “{sval}”)")
                else:
                    rec[key] = d
            elif col.kind == "time":
                t = _coerce_time(raw)
                if t is None:
                    row_errs.append(f"“{col.header}” must be a time like 09:00 (got “{sval}”)")
                else:
                    rec[key] = t
            else:
                rec[key] = sval
        if row_errs:
            errors.append(f"Row {n}: " + "; ".join(row_errs))
        elif any(k in rec for k in by_key):
            out.append(rec)
        if len(errors) >= 25:
            errors.append("…and more. Fix these and re-upload.")
            break
    if errors:
        return [], errors
    if not out:
        return [], ["No data rows found. Fill in at least one row under the headers."]
    return out, []
