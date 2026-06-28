"""One shared Excel look so every Mise download feels like the same professional,
on-brand document: a title band, a calm emerald header row, zebra-striped rows and
soft borders. Import and call style_table() after writing your data."""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Mise emerald palette — calm, professional, peaceful (matches the app brand).
_TITLE = "065F46"      # emerald-900 (title text)
_SUBTLE = "6B7280"     # slate-500 (subtitle)
_HEADER_BG = "047857"  # emerald-700 (header fill)
_HEADER_FG = "FFFFFF"
_ZEBRA = "ECFDF5"      # emerald-50 (alt rows)
_BORDER = "A7F3D0"     # emerald-200
_TOTAL_BG = "D1FAE5"   # emerald-100

_thin = Side(style="thin", color=_BORDER)
_BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def style_table(
    ws: Worksheet,
    *,
    title: str,
    headers: list[str],
    n_rows: int,
    subtitle: str | None = None,
    header_row: int = 3,
    widths: list[int] | None = None,
    right_cols: set[int] | None = None,
) -> None:
    """Apply the Mise look to a simple table on `ws`: a title (+optional subtitle) at
    the top, a coloured header row at `header_row`, then `n_rows` zebra-striped,
    bordered data rows beneath it. Column numbers in `right_cols` are right-aligned."""
    ncols = len(headers)
    right = right_cols or set()

    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=15, color=_TITLE)
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = Font(size=10, italic=True, color=_SUBTLE)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(bold=True, color=_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
        cell.alignment = Alignment(horizontal="right" if c in right else "left", vertical="center")
        cell.border = _BOX
    ws.row_dimensions[header_row].height = 22

    for i in range(n_rows):
        r = header_row + 1 + i
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BOX
            if c in right:
                cell.alignment = Alignment(horizontal="right")
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=_ZEBRA)

    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = (
            widths[c - 1] if widths and c - 1 < len(widths) else 16
        )
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def total_row(
    ws: Worksheet, *, row: int, label: str, label_col: int, value, value_col: int
) -> None:
    """A highlighted totals row (bold emerald on a light fill)."""
    for col, val, right in ((label_col, label, False), (value_col, value, True)):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True, color=_TITLE)
        cell.fill = PatternFill("solid", fgColor=_TOTAL_BG)
        cell.border = _BOX
        if right:
            cell.alignment = Alignment(horizontal="right")
