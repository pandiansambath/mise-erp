"""Build downloadable CSV / Excel / PDF from a P&L dict."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_TITLE = "065F46"      # emerald-900
_HEADER_BG = "047857"  # emerald-700
_KEY_BG = "D1FAE5"     # emerald-100 (key lines)

_PNL_LINES = [
    ("Gross sales", "gross_sales"),
    ("Less: delivery commission", "commission"),
    ("Net sales", "net_sales"),
    ("Cost of sales (food)", "cost_of_sales"),
    ("Gross profit", "gross_profit"),
    ("Operating expenses", "operating_expenses"),
    ("Net profit", "net_profit"),
]
_PNL_PCTS = [
    ("Food cost %", "food_cost_pct"),
    ("Gross margin %", "gross_margin_pct"),
    ("Net margin %", "net_margin_pct"),
]


def to_csv(pnl: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Mise — Profit & Loss"])
    w.writerow(["Period", f"{pnl['date_from']} to {pnl['date_to']}"])
    w.writerow([])
    for label, key in _PNL_LINES:
        w.writerow([label, str(pnl[key])])
    w.writerow([])
    for label, key in _PNL_PCTS:
        w.writerow([label, str(pnl[key])])
    w.writerow([])
    w.writerow(["Expense breakdown"])
    w.writerow(["Category", "Kind", "Total"])
    for c in pnl["expense_breakdown"]:
        w.writerow([c["category_name"], c["kind"], str(c["total"])])
    return buf.getvalue().encode("utf-8")


def to_xlsx(pnl: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"

    title = ws.cell(row=1, column=1, value="Mise — Profit & Loss")
    title.font = Font(bold=True, size=15, color=_TITLE)
    ws.cell(row=2, column=1, value=f"Period: {pnl['date_from']} to {pnl['date_to']}").font = Font(
        size=10, italic=True, color="6B7280"
    )

    r = 4
    for label, key in _PNL_LINES:
        lc = ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=float(pnl[key]))
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right")
        if key in ("net_sales", "gross_profit", "net_profit"):
            fill = PatternFill("solid", fgColor=_KEY_BG)
            lc.font = Font(bold=True, color=_TITLE)
            lc.fill = fill
            cell.font = Font(bold=True, color=_TITLE)
            cell.fill = fill
        r += 1
    r += 1
    for label, key in _PNL_PCTS:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=float(pnl[key]))
        r += 1

    r += 1
    hdr = ws.cell(row=r, column=1, value="Expense breakdown")
    hdr.font = Font(bold=True, size=12, color=_TITLE)
    r += 1
    for col, name in enumerate(["Category", "Kind", "Total"], start=1):
        c = ws.cell(row=r, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_HEADER_BG)
    r += 1
    for c in pnl["expense_breakdown"]:
        ws.cell(row=r, column=1, value=c["category_name"])
        ws.cell(row=r, column=2, value=c["kind"])
        ws.cell(row=r, column=3, value=float(c["total"])).number_format = "#,##0.00"
        r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def to_pdf(pnl: dict) -> bytes:
    """One-page branded P&L snapshot — what the monthly archive hands back."""
    from fpdf.enums import XPos, YPos

    from app.core.pdf import branded_pdf, footer, ps

    pdf = branded_pdf("Profit & Loss", f"{pnl['date_from']} to {pnl['date_to']}")

    def line(label: str, value: str, *, bold: bool = False, gap: float = 0.0) -> None:
        if gap:
            pdf.set_y(pdf.get_y() + gap)
        pdf.set_x(14)
        pdf.set_font("Helvetica", "B" if bold else "", 10)
        pdf.cell(120, 7.5, text=ps(label))
        pdf.cell(60, 7.5, text=ps(value), align="R", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    for label, key in _PNL_LINES:
        line(label, str(pnl[key]), bold=key in ("net_sales", "gross_profit", "net_profit"))
    for label, key in _PNL_PCTS:
        line(label, f"{pnl[key]}%")
    if pnl.get("waste_total") is not None:
        line("Logged waste", str(pnl["waste_total"]))

    if pnl["expense_breakdown"]:
        pdf.set_y(pdf.get_y() + 4)
        pdf.set_x(14)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, text="Expense breakdown", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font("Helvetica", "", 9)
        for c in pnl["expense_breakdown"]:
            pdf.set_x(14)
            pdf.cell(100, 6.5, text=ps(c["category_name"]))
            pdf.cell(30, 6.5, text=ps(c["kind"].lower()), align="C")
            pdf.cell(
                50, 6.5, text=ps(str(c["total"])), align="R",
                new_x=XPos.LMARGIN, new_y=YPos.NEXT,
            )

    footer(pdf)
    return bytes(pdf.output())
